#!/usr/bin/env python3
"""
Excel Template Integration System (Simplified)
Handles bid template processing without pandas dependency
"""

import os
import json
import tempfile
from datetime import datetime
from typing import Dict, List, Optional
import logging
from openpyxl import load_workbook, Workbook

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelTemplateProcessor:
    """Processes Excel templates for bid submissions and comparisons"""
    
    def __init__(self, templates_dir: str = "templates"):
        self.templates_dir = templates_dir
        self.bid_templates = {}
        self.comparison_templates = {}
        self.load_templates()
    
    def load_templates(self):
        """Load available Excel templates"""
        try:
            # Load bid templates
            bid_template_path = os.path.join(self.templates_dir, "teklif_formu.xlsx")
            if os.path.exists(bid_template_path):
                self.bid_templates["standard"] = bid_template_path
                logger.info(f"Loaded bid template: {bid_template_path}")
            
            # Load comparison templates
            comparison_template_path = os.path.join(self.templates_dir, "teklif_mukayese_formu.xlsx")
            if os.path.exists(comparison_template_path):
                self.comparison_templates["standard"] = comparison_template_path
                logger.info(f"Loaded comparison template: {comparison_template_path}")
                
        except Exception as e:
            logger.error(f"Error loading templates: {e}")
    
    def generate_bid_template(self, demand_data: Dict, supplier_data: Dict) -> str:
        """
        Generate bid template for supplier
        
        Args:
            demand_data: Demand information from Firestore
            supplier_data: Supplier information
            
        Returns:
            Path to generated Excel file
        """
        try:
            template_path = self.bid_templates.get("standard")
            if not template_path:
                # Create a simple template if none exists
                return self.create_simple_bid_template(demand_data, supplier_data)
            
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{demand_data.get('satfk', 'UNKNOWN')}_Teklif_{supplier_data.get('companyName', 'Supplier')}_{timestamp}.xlsx"
            output_path = os.path.join("generated", output_filename)
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Load template
            workbook = load_workbook(template_path)
            
            # Fill demand information
            if "Talep_Bilgileri" in workbook.sheetnames:
                sheet = workbook["Talep_Bilgileri"]
                self.fill_demand_info_sheet(sheet, demand_data)
            
            # Fill supplier information
            if "Tedarikci_Bilgileri" in workbook.sheetnames:
                sheet = workbook["Tedarikci_Bilgileri"]
                self.fill_supplier_info_sheet(sheet, supplier_data)
            
            # Save filled template
            workbook.save(output_path)
            
            logger.info(f"Generated bid template: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating bid template: {e}")
            # Fallback to simple template
            return self.create_simple_bid_template(demand_data, supplier_data)
    
    def create_simple_bid_template(self, demand_data: Dict, supplier_data: Dict) -> str:
        """Create a simple bid template if no template exists"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{demand_data.get('satfk', 'UNKNOWN')}_Teklif_{supplier_data.get('companyName', 'Supplier')}_{timestamp}.xlsx"
            output_path = os.path.join("generated", output_filename)
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Create new workbook
            workbook = Workbook()
            
            # Create demand info sheet
            demand_sheet = workbook.active
            demand_sheet.title = "Talep_Bilgileri"
            
            # Add demand information
            demand_info = [
                ["SATFK", demand_data.get("satfk", "")],
                ["Talep Tarihi", demand_data.get("demandDate", "")],
                ["Termin Tarihi", demand_data.get("dueDate", "")],
                ["Talep Başlığı", demand_data.get("title", "")],
                ["Öncelik", demand_data.get("priority", "")],
                ["Para Birimi", demand_data.get("currency", "TRY")],
                ["Kategoriler", ", ".join(demand_data.get("categoryTags", []))],
                ["Açıklama", demand_data.get("spec", "")],
                ["Teslimat Adresi", demand_data.get("deliveryAddress", "")],
                ["Ödeme Koşulları", demand_data.get("paymentTerms", "")]
            ]
            
            for row, (label, value) in enumerate(demand_info, 1):
                demand_sheet.cell(row=row, column=1, value=label)
                demand_sheet.cell(row=row, column=2, value=value)
            
            # Create supplier info sheet
            supplier_sheet = workbook.create_sheet("Tedarikci_Bilgileri")
            
            supplier_info = [
                ["Şirket Adı", supplier_data.get("companyName", "")],
                ["Vergi No", supplier_data.get("taxNumber", "")],
                ["MERSİS No", supplier_data.get("mersisNo", "")],
                ["Adres", supplier_data.get("address", "")],
                ["E-posta", supplier_data.get("email", "")],
                ["Telefon", ", ".join(supplier_data.get("contactPhones", []))],
                ["Kategoriler", ", ".join(supplier_data.get("categories", []))],
                ["Web Sitesi", supplier_data.get("website", "")]
            ]
            
            for row, (label, value) in enumerate(supplier_info, 1):
                supplier_sheet.cell(row=row, column=1, value=label)
                supplier_sheet.cell(row=row, column=2, value=value)
            
            # Create bid items sheet
            items_sheet = workbook.create_sheet("Teklif_Kalemleri")
            
            # Add headers
            headers = ["Sıra No", "Malzeme Kodu", "Tanım", "Marka/Model", "Miktar", "Birim", 
                      "Birim Fiyat", "Toplam Fiyat", "Teslim Tarihi", "Garanti", "Menşei", "Notlar"]
            
            for col, header in enumerate(headers, 1):
                items_sheet.cell(row=1, column=col, value=header)
            
            # Add sample row
            sample_data = ["1", "", "", "", "1", "adet", "", "", "", "", "", ""]
            for col, value in enumerate(sample_data, 1):
                items_sheet.cell(row=2, column=col, value=value)
            
            # Save workbook
            workbook.save(output_path)
            
            logger.info(f"Created simple bid template: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error creating simple bid template: {e}")
            raise
    
    def fill_demand_info_sheet(self, sheet, demand_data: Dict):
        """Fill demand information in template sheet"""
        try:
            # Map demand fields to template cells (adjust based on your template)
            field_mapping = {
                "B2": demand_data.get("satfk", ""),  # SATFK
                "B3": demand_data.get("demandDate", ""),  # Talep Tarihi
                "B4": demand_data.get("dueDate", ""),  # Termin Tarihi
                "B5": demand_data.get("title", ""),  # Talep Başlığı
                "B6": demand_data.get("priority", ""),  # Öncelik
                "B7": demand_data.get("currency", "TRY"),  # Para Birimi
                "B8": ", ".join(demand_data.get("categoryTags", [])),  # Kategoriler
                "B9": demand_data.get("spec", ""),  # Açıklama
                "B10": demand_data.get("deliveryAddress", ""),  # Teslimat Adresi
                "B11": demand_data.get("paymentTerms", "")  # Ödeme Koşulları
            }
            
            # Update sheet with demand data
            for cell_ref, value in field_mapping.items():
                if cell_ref in sheet:
                    sheet[cell_ref] = value
            
        except Exception as e:
            logger.error(f"Error filling demand info: {e}")
    
    def fill_supplier_info_sheet(self, sheet, supplier_data: Dict):
        """Fill supplier information in template sheet"""
        try:
            # Map supplier fields to template cells (adjust based on your template)
            field_mapping = {
                "B2": supplier_data.get("companyName", ""),  # Şirket Adı
                "B3": supplier_data.get("taxNumber", ""),  # Vergi No
                "B4": supplier_data.get("mersisNo", ""),  # MERSİS No
                "B5": supplier_data.get("address", ""),  # Adres
                "B6": supplier_data.get("email", ""),  # E-posta
                "B7": ", ".join(supplier_data.get("contactPhones", [])),  # Telefon
                "B8": ", ".join(supplier_data.get("categories", [])),  # Kategoriler
                "B9": supplier_data.get("website", "")  # Web Sitesi
            }
            
            # Update sheet with supplier data
            for cell_ref, value in field_mapping.items():
                if cell_ref in sheet:
                    sheet[cell_ref] = value
            
        except Exception as e:
            logger.error(f"Error filling supplier info: {e}")
    
    def process_submitted_bid(self, bid_file_path: str, demand_data: Dict) -> Dict:
        """
        Process submitted bid Excel file and extract bid information
        
        Args:
            bid_file_path: Path to submitted bid Excel file
            demand_data: Original demand data
            
        Returns:
            Extracted bid data as dictionary
        """
        try:
            # Load submitted bid
            workbook = load_workbook(bid_file_path)
            
            bid_data = {
                "submitted_at": datetime.now().isoformat(),
                "file_path": bid_file_path,
                "demand_id": demand_data.get("id"),
                "satfk": demand_data.get("satfk"),
                "items": [],
                "commercial_terms": {},
                "total_amount": 0
            }
            
            # Extract bid items
            if "Teklif_Kalemleri" in workbook.sheetnames:
                items_sheet = workbook["Teklif_Kalemleri"]
                bid_data["items"] = self.extract_bid_items_sheet(items_sheet)
            
            # Extract commercial terms
            if "Ticari_Sartlar" in workbook.sheetnames:
                terms_sheet = workbook["Ticari_Sartlar"]
                bid_data["commercial_terms"] = self.extract_commercial_terms_sheet(terms_sheet)
            
            # Calculate total amount
            bid_data["total_amount"] = sum(item.get("total_price", 0) for item in bid_data["items"])
            
            logger.info(f"Processed bid file: {bid_file_path}")
            return bid_data
            
        except Exception as e:
            logger.error(f"Error processing bid file: {e}")
            raise
    
    def extract_bid_items_sheet(self, sheet) -> List[Dict]:
        """Extract bid items from Excel sheet"""
        items = []
        
        try:
            # Skip header row (row 1)
            for row in range(2, sheet.max_row + 1):
                material_code = sheet.cell(row=row, column=2).value  # Malzeme Kodu
                
                if material_code:  # Skip empty rows
                    item = {
                        "material_code": str(material_code),
                        "description": str(sheet.cell(row=row, column=3).value or ""),
                        "brand_model": str(sheet.cell(row=row, column=4).value or ""),
                        "quantity": float(sheet.cell(row=row, column=5).value or 0),
                        "unit": str(sheet.cell(row=row, column=6).value or ""),
                        "unit_price": float(sheet.cell(row=row, column=7).value or 0),
                        "total_price": float(sheet.cell(row=row, column=8).value or 0),
                        "delivery_date": str(sheet.cell(row=row, column=9).value or ""),
                        "warranty": str(sheet.cell(row=row, column=10).value or ""),
                        "origin": str(sheet.cell(row=row, column=11).value or ""),
                        "notes": str(sheet.cell(row=row, column=12).value or "")
                    }
                    items.append(item)
            
            return items
            
        except Exception as e:
            logger.error(f"Error extracting bid items: {e}")
            return []
    
    def extract_commercial_terms_sheet(self, sheet) -> Dict:
        """Extract commercial terms from Excel sheet"""
        terms = {}
        
        try:
            for row in range(1, sheet.max_row + 1):
                condition = sheet.cell(row=row, column=1).value
                value = sheet.cell(row=row, column=2).value
                
                if condition and value:
                    terms[str(condition)] = str(value)
            
            return terms
            
        except Exception as e:
            logger.error(f"Error extracting commercial terms: {e}")
            return {}
    
    def generate_comparison_table(self, demand_data: Dict, bids_data: List[Dict], 
                                user_premium: bool = False) -> str:
        """
        Generate comparison table for received bids
        
        Args:
            demand_data: Original demand data
            bids_data: List of processed bid data
            user_premium: Whether user has premium membership
            
        Returns:
            Path to generated comparison Excel file
        """
        try:
            # Limit bids for non-premium users
            if not user_premium and len(bids_data) > 3:
                bids_data = bids_data[:3]
                logger.info(f"Limited to 3 bids for non-premium user")
            
            template_path = self.comparison_templates.get("standard")
            if not template_path:
                # Create a simple comparison template
                return self.create_simple_comparison_template(demand_data, bids_data)
            
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{demand_data.get('satfk', 'UNKNOWN')}_Mukayese_{timestamp}.xlsx"
            output_path = os.path.join("generated", output_filename)
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Load template
            workbook = load_workbook(template_path)
            
            # Fill comparison data
            if "Mukayese_Tablosu" in workbook.sheetnames:
                comparison_sheet = workbook["Mukayese_Tablosu"]
                self.fill_comparison_data_sheet(comparison_sheet, demand_data, bids_data)
            
            # Save comparison table
            workbook.save(output_path)
            
            logger.info(f"Generated comparison table: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating comparison table: {e}")
            # Fallback to simple template
            return self.create_simple_comparison_template(demand_data, bids_data)
    
    def create_simple_comparison_template(self, demand_data: Dict, bids_data: List[Dict]) -> str:
        """Create a simple comparison template if no template exists"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{demand_data.get('satfk', 'UNKNOWN')}_Mukayese_{timestamp}.xlsx"
            output_path = os.path.join("generated", output_filename)
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Create new workbook
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Mukayese_Tablosu"
            
            # Add demand information header
            sheet.cell(row=1, column=1, value=f"SATFK: {demand_data.get('satfk', '')}")
            sheet.cell(row=2, column=1, value=f"Talep Başlığı: {demand_data.get('title', '')}")
            sheet.cell(row=3, column=1, value=f"Termin Tarihi: {demand_data.get('dueDate', '')}")
            
            # Add bid columns dynamically
            start_col = 2
            for i, bid in enumerate(bids_data):
                col_idx = start_col + i
                
                # Supplier name
                sheet.cell(row=4, column=col_idx, value=bid.get("supplier_name", f"Tedarikçi {i+1}"))
                
                # Total amount
                sheet.cell(row=5, column=col_idx, value=bid.get("total_amount", 0))
                
                # Commercial terms
                terms = bid.get("commercial_terms", {})
                row_idx = 6
                for term, value in terms.items():
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                    row_idx += 1
            
            # Save workbook
            workbook.save(output_path)
            
            logger.info(f"Created simple comparison template: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error creating simple comparison template: {e}")
            raise
    
    def fill_comparison_data_sheet(self, sheet, demand_data: Dict, bids_data: List[Dict]):
        """Fill comparison table with bid data"""
        try:
            # Add demand information header
            sheet.cell(row=1, column=1, value=f"SATFK: {demand_data.get('satfk', '')}")
            sheet.cell(row=2, column=1, value=f"Talep Başlığı: {demand_data.get('title', '')}")
            sheet.cell(row=3, column=1, value=f"Termin Tarihi: {demand_data.get('dueDate', '')}")
            
            # Add bid columns dynamically
            start_col = 2
            for i, bid in enumerate(bids_data):
                col_idx = start_col + i
                
                # Supplier name
                sheet.cell(row=4, column=col_idx, value=bid.get("supplier_name", f"Tedarikçi {i+1}"))
                
                # Total amount
                sheet.cell(row=5, column=col_idx, value=bid.get("total_amount", 0))
                
                # Commercial terms
                terms = bid.get("commercial_terms", {})
                row_idx = 6
                for term, value in terms.items():
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                    row_idx += 1
            
        except Exception as e:
            logger.error(f"Error filling comparison data: {e}")

def main():
    """Test the Excel template processor"""
    processor = ExcelTemplateProcessor()
    
    # Sample demand data
    demand_data = {
        "id": "test_demand_123",
        "satfk": "SATFK-20251024-00C9",
        "title": "Test Talep",
        "demandDate": "2025-10-24",
        "dueDate": "2025-10-26",
        "priority": "price",
        "currency": "TRY",
        "categoryTags": ["makine-imalat", "elektrik"],
        "spec": "Test açıklama",
        "deliveryAddress": "Test adres",
        "paymentTerms": "30 gün vadeli"
    }
    
    # Sample supplier data
    supplier_data = {
        "companyName": "Test Tedarikçi A.Ş.",
        "taxNumber": "1234567890",
        "mersisNo": "0123456789012345",
        "address": "Test adres",
        "email": "test@supplier.com",
        "contactPhones": ["+90 555 123 4567"],
        "categories": ["makine-imalat"],
        "website": "https://test-supplier.com"
    }
    
    try:
        # Generate bid template
        bid_template_path = processor.generate_bid_template(demand_data, supplier_data)
        print(f"Generated bid template: {bid_template_path}")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()