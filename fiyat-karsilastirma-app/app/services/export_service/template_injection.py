"""
Template injection export service using openpyxl
Loads TEKLİF MUKAYESE FORMU.xlsx template and injects data
"""
import os
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None
    load_workbook = None

from .base import BaseExportService
from .template_mapping import TEMPLATE_MAPPING, get_vendor_column_mapping, get_footer_row_mapping
from ...domain.models import ComparisonResult, Vendor, ComparisonRow


class TemplateInjectionExportService(BaseExportService):
    """Template injection ile Excel export"""
    
    def __init__(self, comparison_result: ComparisonResult):
        super().__init__(comparison_result)
        self.template_path = self._get_template_path()
    
    def _get_template_path(self) -> str:
        """Template dosya yolunu getir"""
        # Proje root'undan template path'i oluştur
        current_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
        template_path = os.path.join(current_dir, "assets", "TEKLİF MUKAYESE FORMU.xlsx")
        return template_path
    
    def _validate_template_exists(self) -> bool:
        """Template dosyasının varlığını kontrol et"""
        return os.path.exists(self.template_path)
    
    def export_to_excel(self, output_path: str) -> str:
        """Template injection ile Excel export"""
        if not openpyxl:
            raise ImportError("openpyxl kütüphanesi yüklü değil. pip install openpyxl")
        
        if not self._validate_template_exists():
            raise FileNotFoundError(
                f"Şablon dosyası eksik: {self.template_path}\n"
                f"Lütfen TEKLİF MUKAYESE FORMU.xlsx dosyasını assets/ klasörüne koyun."
            )
        
        # Output path'i validate et
        self._validate_output_path(output_path)
        
        # Dosya adını oluştur
        filename = self._get_timestamped_filename("xlsx")
        full_path = os.path.join(output_path, filename)
        
        # Template'i yükle
        workbook = load_workbook(self.template_path)
        
        # Vendor gruplarını al
        vendor_groups = self._get_vendor_groups_for_export()
        
        # Her vendor grubu için sheet oluştur
        for i, vendor_group in enumerate(vendor_groups):
            if i == 0:
                # İlk grup için mevcut sheet'i kullan
                worksheet = workbook.active
                if len(vendor_groups) > 1:
                    worksheet.title = self._get_sheet_name_for_vendor_group(vendor_group, i)
            else:
                # Diğer gruplar için yeni sheet oluştur
                base_sheet = workbook.active
                worksheet = workbook.copy_worksheet(base_sheet)
                worksheet.title = self._get_sheet_name_for_vendor_group(vendor_group, i)
            
            # Bu sheet'e veri yaz
            self._write_data_to_worksheet(worksheet, vendor_group)
        
        # Workbook'u kaydet
        workbook.save(full_path)
        workbook.close()
        
        return full_path
    
    def export_to_csv(self, output_path: str) -> str:
        """CSV export (base class'tan inherit)"""
        filename = self._get_timestamped_filename("csv")
        full_path = os.path.join(output_path, filename)
        
        # CSV verisini hazırla
        csv_data = self._prepare_csv_data()
        
        # CSV dosyasını yaz
        with open(full_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            if csv_data:
                writer = csv.DictWriter(csvfile, fieldnames=csv_data[0].keys())
                writer.writeheader()
                writer.writerows(csv_data)
        
        return full_path
    
    def _write_data_to_worksheet(self, worksheet, vendor_group: List[Vendor]):
        """Worksheet'e veri yaz"""
        # Mapping'i al
        mapping = TEMPLATE_MAPPING
        data_start_row = mapping["data_start_row"]
        
        # Her ürün için satır yaz
        for i, comparison_row in enumerate(self.comparison_result.comparison_rows):
            row_num = data_start_row + i
            
            # Temel bilgileri yaz
            self._write_basic_data(worksheet, row_num, comparison_row)
            
            # Vendor bilgilerini yaz
            self._write_vendor_data(worksheet, row_num, comparison_row, vendor_group)
        
        # Footer bilgilerini yaz
        self._write_footer_data(worksheet, vendor_group)
    
    def _write_basic_data(self, worksheet, row_num: int, comparison_row: ComparisonRow):
        """Temel bilgileri yaz (NO, HİZMETİN ADI, MİKTAR, BİRİM)"""
        mapping = TEMPLATE_MAPPING["columns"]
        
        worksheet[f"{mapping['no']}{row_num}"] = comparison_row.no
        worksheet[f"{mapping['hizmet_adi']}{row_num}"] = comparison_row.hizmet_adi
        worksheet[f"{mapping['miktar']}{row_num}"] = comparison_row.miktar
        worksheet[f"{mapping['birim']}{row_num}"] = comparison_row.birim
    
    def _write_vendor_data(self, worksheet, row_num: int, comparison_row: ComparisonRow, vendor_group: List[Vendor]):
        """Vendor bilgilerini yaz"""
        for i, vendor in enumerate(vendor_group):
            if i >= len(TEMPLATE_MAPPING["columns"]["vendor"]):
                break  # Max 5 vendor desteği
            
            vendor_mapping = get_vendor_column_mapping(i)
            offer = comparison_row.vendor_offers.get(vendor.firma_adi)
            
            if offer:
                # Teklif var
                worksheet[f"{vendor_mapping['birim_fiyat']}{row_num}"] = offer.birim_fiyat
                worksheet[f"{vendor_mapping['toplam']}{row_num}"] = offer.toplam
                worksheet[f"{vendor_mapping['toplam_tl']}{row_num}"] = offer.toplam_tl
            else:
                # Teklif yok, boş bırak
                worksheet[f"{vendor_mapping['birim_fiyat']}{row_num}"] = None
                worksheet[f"{vendor_mapping['toplam']}{row_num}"] = None
                worksheet[f"{vendor_mapping['toplam_tl']}{row_num}"] = None
    
    def _write_footer_data(self, worksheet, vendor_group: List[Vendor]):
        """Footer bilgilerini yaz"""
        footer_mapping = get_footer_row_mapping()
        footer_col = footer_mapping["footer_text_column"]
        
        # En iyi tedarikçiyi bul
        best_vendor, best_total = self._get_best_overall_vendor()
        
        # Footer metinlerini yaz
        if best_vendor and best_total:
            # En uygun firma bilgisini yaz
            footer_text = f"Teknik Değerlendirme sonucunda en uygun bulunan firma: {best_vendor} (Toplam: {best_total:,.2f} TL)"
        else:
            footer_text = "Teknik Değerlendirme sonucunda en uygun bulunan firma: Değerlendirme yapılamadı"
        
        worksheet[f"{footer_col}{footer_mapping['en_uygun_firma_satir']}"] = footer_text
        
        # Diğer footer alanları (opsiyonel)
        # ÖDEME, TESLİM, NOT alanları için boş bırakabiliriz veya varsayılan değerler yazabiliriz
        worksheet[f"{footer_col}{footer_mapping['odeme_satir']}"] = "Ödeme şekli: Belirtilmedi"
        worksheet[f"{footer_col}{footer_mapping['teslim_satir']}"] = "Teslim şekli: Belirtilmedi"
        worksheet[f"{footer_col}{footer_mapping['not_satir']}"] = "Not: Belirtilmedi"
    
    def get_template_info(self) -> Dict[str, Any]:
        """Template dosyası hakkında bilgi getir"""
        return {
            "template_path": self.template_path,
            "template_exists": self._validate_template_exists(),
            "mapping": TEMPLATE_MAPPING,
            "max_vendors_per_sheet": TEMPLATE_MAPPING["columns"]["vendor"].__len__(),
            "supported_export_modes": ["template_injection"]
        }
