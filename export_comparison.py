"""
Teklif Mukayese Çıktısı (JSON + teklifler -> Excel)
C:\Users\faruk\OneDrive\Desktop\teklifbul-web\assets\teklif mukayese formu konumundaki tablo.

Amaç:
- Talep JSON'u (talep.json) ve teklifler (bids.json) ile
  'teklif mukayese formu.xlsx' şablonunu bozulmadan doldur.
- Ürün sırasını aynen koru; birim fiyat * miktar = satır toplam.
- F21/I21/L21/... hücrelerine firma toplamını yaz; ödeme/termin/teslim/açıklama alanlarını doldur.
- Premium (3+ firma) durumunda 3'er sütun bloklarını sağa doğru otomatik genişlet.

Gereksinim:
  pip install openpyxl

Beklenen dosyalar:
  ./talep.json   (Prompt #1'in çıktısı)
  ./bids.json    (aşağıdaki şema)
  ./teklif mukayese formu.xlsx

bids.json şeması:
{
  "firms": [
    {
      "firma_adi": "Tedarikçi A",
      "tel": "0500...",
      "odeme": "%50 Peşin",
      "termin": "2025-10-25",
      "fiyatlar": [ 12.5, 8.9, null, 33.0 ]  // talep.items sırasına göre; fiyat yoksa null
    },
    ...
  ]
}
"""

import json
import sys
import io
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

COMPARE_TOP_MAP = {
    "talep_kodu": "T1",
    "stf_no": "T2",
    "santiye": "T3",
    "usd_try": "T4",
    "talep_konusu": "B5",
}

START_ROW_ITEMS = 8
TOTAL_ROW = 21
PAY_ROW = 22
TERM_ROW = 23
DELIV_ROW = 24
NOTE_ROW = 25

# İlk 5 firma için başlangıç kolonları (3'er sütun blok mantığına uyumlu olacak şekilde: F/G, I/J, L/M, O/P, R/S)
BASE_COLS = ["F", "I", "L", "O", "R"]  # gerekirse otomatik büyütülecek

# === STİL TANIMLARI (Şablona Uygun) ===
# Kenarlık: İnce siyah çizgi
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Arka plan renkleri
LIGHT_GRAY_FILL = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# Yazı stilleri
BOLD_FONT = Font(name='Calibri', size=11, bold=True)
NORMAL_FONT = Font(name='Calibri', size=11, bold=False)
TITLE_FONT = Font(name='Calibri', size=12, bold=True)

# Hizalama
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
LEFT_ALIGNMENT = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGNMENT = Alignment(horizontal='right', vertical='center')

def _col_letter(idx: int) -> str:
    return get_column_letter(idx)

def _ensure_col(fi: int, base_cols: List[str]) -> str:
    """Firma index'i için kullanılacak baz kolonu (birim fiyat sütunu) döndürür."""
    if fi < len(base_cols):
        return base_cols[fi]
    # Dinamik genişletme: son kolonun indexinden itibaren 3 sütunluk bloklara devam
    last = base_cols[-1]
    base_idx = column_index_from_string(last)
    add = (fi - len(base_cols) + 1) * 3
    new_idx = base_idx + add
    return _col_letter(new_idx)

def load_json(path: str) -> Dict[str,Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def apply_cell_style(ws, cell_address, font=None, fill=None, border=None, alignment=None):
    """Hücreye stil uygula"""
    cell = ws[cell_address]
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment

def merge_and_style_cells(ws, range_str, value, font=None, fill=None, border=None, alignment=None):
    """Hücreleri birleştir ve stil uygula"""
    ws.merge_cells(range_str)
    apply_cell_style(ws, range_str.split(':')[0], font, fill, border, alignment)
    ws[range_str.split(':')[0]] = value

def firm_total(firm: Dict[str,Any], items: List[Dict[str,Any]]) -> float:
    total = 0.0
    for i, it in enumerate(items):
        price = firm.get("fiyatlar", [None]*len(items))[i] if i < len(firm.get("fiyatlar", [])) else None
        if price is not None:
            total += float(price) * float(it.get("qty") or 0)
    return total

def main(talep_json="talep.json", bids_json="bids.json", template="assets/teklif mukayese formu.xlsx", out_path=None):
    try:
        print(f"📊 Talep JSON'u okunuyor: {talep_json}")
        talep = load_json(talep_json)
        
        print(f"📊 Teklifler JSON'u okunuyor: {bids_json}")
        bids = load_json(bids_json)
        
        print(f"📊 Excel şablonu okunuyor: {template}")
        wb = openpyxl.load_workbook(template)
        ws = wb.active
        
        items = sorted(talep["items"], key=lambda x: x.get("no", 0))
        print(f"✅ {len(items)} adet ürün kalemi bulundu")

        # === 1. BAŞLIK VE ÜST BİLGİLER ===
        print("📋 Başlık ve üst bilgiler yazılıyor...")
        
        # Ana başlık: FİYAT KARŞILAŞTIRMA TABLOSU (F2:Q2)
        merge_and_style_cells(ws, "F2:Q2", "FİYAT KARŞILAŞTIRMA TABLOSU", 
                             TITLE_FONT, LIGHT_GRAY_FILL, THIN_BORDER, CENTER_ALIGNMENT)
        
        # Alt başlık: FİRMA İSİMLERİ (F5:Q5)
        merge_and_style_cells(ws, "F5:Q5", "FİRMA İSİMLERİ", 
                             BOLD_FONT, LIGHT_GRAY_FILL, THIN_BORDER, CENTER_ALIGNMENT)
        
        # Talep bilgileri (S1-T4)
        demand_labels = ["Talep Kodu", "STF No", "Şantiye", "DOLAR"]
        for i, label in enumerate(demand_labels):
            row = i + 1
            apply_cell_style(ws, f"S{row}", BOLD_FONT, LIGHT_GRAY_FILL, THIN_BORDER, LEFT_ALIGNMENT)
            ws[f"S{row}"] = label
            
            apply_cell_style(ws, f"T{row}", NORMAL_FONT, WHITE_FILL, THIN_BORDER, LEFT_ALIGNMENT)
            if row == 1 and "talep_kodu" in talep:
                ws[f"T{row}"] = talep["talep_kodu"]
            elif row == 2 and "stf_no" in talep:
                ws[f"T{row}"] = talep["stf_no"]
            elif row == 3 and "santiye" in talep:
                ws[f"T{row}"] = talep["santiye"]
            elif row == 4 and "usd_try" in talep:
                ws[f"T{row}"] = talep["usd_try"]
        
        # Konu bilgisi (A5-B5)
        apply_cell_style(ws, "A5", BOLD_FONT, LIGHT_GRAY_FILL, THIN_BORDER, LEFT_ALIGNMENT)
        ws["A5"] = "KONU"
        apply_cell_style(ws, "B5", NORMAL_FONT, WHITE_FILL, THIN_BORDER, LEFT_ALIGNMENT)
        ws["B5"] = talep.get("talep_konusu", "")

        # === 2. TABLO BAŞLIKLARI ===
        print("📊 Tablo başlıkları yazılıyor...")
        
        # Sol sütun başlıkları
        headers = ["NO", "ÜRÜN İSMİ", "MİKTAR", "BİRİM"]
        for i, header in enumerate(headers):
            col = chr(ord('B') + i)
            apply_cell_style(ws, f"{col}6", BOLD_FONT, LIGHT_GRAY_FILL, THIN_BORDER, CENTER_ALIGNMENT)
            ws[f"{col}6"] = header
        
        # === 3. ÜRÜN KALEMLERİ ===
        print("📦 Ürün kalemleri yazılıyor...")
        for i, it in enumerate(items):
            r = START_ROW_ITEMS + i
            # Sıra no
            apply_cell_style(ws, f"B{r}", NORMAL_FONT, LIGHT_GRAY_FILL, THIN_BORDER, CENTER_ALIGNMENT)
            ws[f"B{r}"] = it.get("no", i+1)
            # Ürün adı
            apply_cell_style(ws, f"C{r}", NORMAL_FONT, WHITE_FILL, THIN_BORDER, LEFT_ALIGNMENT)
            ws[f"C{r}"] = it.get("name", "")
            # Miktar
            apply_cell_style(ws, f"D{r}", NORMAL_FONT, WHITE_FILL, THIN_BORDER, CENTER_ALIGNMENT)
            ws[f"D{r}"] = float(it.get("qty") or 0)
            # Birim
            apply_cell_style(ws, f"E{r}", NORMAL_FONT, WHITE_FILL, THIN_BORDER, CENTER_ALIGNMENT)
            ws[f"E{r}"] = it.get("unit", "")
            print(f"  Satır {r}: {it.get('name', '')} - {it.get('qty', 0)} {it.get('unit', '')}")

        # Firmaları toplamlarına göre sırala (en ucuz en solda)
        firms_sorted = sorted(bids.get("firms", []), key=lambda f: firm_total(f, items))
        print(f"🏢 {len(firms_sorted)} adet firma teklifi işleniyor...")

        for fi, firm in enumerate(firms_sorted):
            col = _ensure_col(fi, BASE_COLS)
            col_idx = column_index_from_string(col)
            col_total = 0.0

            # Firma başlığı (F6/I6/...)
            title = firm.get("firma_adi","")
            tel = firm.get("tel")
            ws[f"{col}6"] = f"{title} {('('+tel+')') if tel else ''}"
            print(f"  {col}6: {title} ({tel})")

            # Satırlar: birim fiyat + yan sütunda satır toplamı
            for i, it in enumerate(items):
                r = START_ROW_ITEMS + i
                price = None
                if i < len(firm.get("fiyatlar", [])):
                    price = firm["fiyatlar"][i]
                if price is None:
                    continue

                ws[f"{col}{r}"] = float(price)  # birim fiyat (F8, F9...)
                next_col = get_column_letter(col_idx + 1)  # G8, G9...
                line_total = float(price) * float(it.get("qty") or 0)
                ws[f"{next_col}{r}"] = line_total
                col_total += line_total
                print(f"    {col}{r}: {price} -> {next_col}{r}: {line_total}")

            # Alt satırlar
            ws[f"{col}{TOTAL_ROW}"] = col_total
            ws[f"{col}{PAY_ROW}"]   = firm.get("odeme", "")
            ws[f"{col}{TERM_ROW}"]  = firm.get("termin", "")
            ws[f"{col}{DELIV_ROW}"] = talep.get("alim_yeri", "")
            ws[f"{col}{NOTE_ROW}"]  = talep.get("aciklama", "")
            print(f"    {col}{TOTAL_ROW}: Toplam = {col_total}")

        # Kaydet
        out_name = out_path or f"mukayese_tablo_{(talep.get('stf_no') or 'STF')}.xlsx"
        wb.save(out_name)
        print(f"✅ Excel oluşturuldu: {out_name}")
        
    except FileNotFoundError as e:
        print(f"❌ Dosya bulunamadı: {e}")
        print("📁 Gerekli dosyalar:")
        print("  - talep.json (Prompt #1'in çıktısı)")
        print("  - bids.json (teklif verileri)")
        print("  - assets/teklif mukayese formu.xlsx (şablon)")
    except Exception as e:
        print(f"❌ Hata oluştu: {e}")
        print("🔧 JSON formatını ve şablon hücrelerini kontrol edin")

if __name__ == "__main__":
    # python export_comparison.py
    # veya: python export_comparison.py talep.json bids.json "assets/teklif mukayese formu.xlsx" "cikti.xlsx"
    if len(sys.argv) >= 2:
        args = sys.argv[1:]
        main(*args)
    else:
        main()
