"""
Form parser service for satın alma talep formu.xlsx
Parses Excel template and creates PurchaseRequest objects
"""
import os
import pandas as pd
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None
    load_workbook = None

from ..domain.models import PurchaseRequest, PurchaseItem


# Excel form mapping constants
FORM_MAP = {
    "sheet_name": "SATINALMA",
    "header_row": 7,
    "data_start_row": 9,
    "columns": {
        "sira_no": "A",
        "malzeme_kodu": "B",
        "malzeme_tanimi": "C",
        "marka": "D",
        "birim": "E",
        "miktar": "F",
        "istenilen_teslim_tarihi": "G",
        "ambardaki_miktar": "H",
        "siparis_miktari": "I",
        "hedef_fiyat": "J",
        "genel_toplam_tl": "L"
    },
    "meta_rows": {
        "santiye": 3,
        "stf_no": 4,
        "stf_tarihi": 5,
        "alim_yeri": 6
    },
    "meta_value_search_right_from_col": 8  # scan cols 9..13 same row for first non-empty
}


class FormParserService:
    """Satın alma talep formu parser servisi"""
    
    def __init__(self, template_path: str = None):
        if template_path is None:
            # Proje root'undan template path'i oluştur
            current_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            template_path = os.path.join(current_dir, "assets", "satın alma talep formu.xlsx")
        
        self.template_path = template_path
    
    def validate_template_exists(self) -> bool:
        """Template dosyasının varlığını kontrol et"""
        return os.path.exists(self.template_path)
    
    def download_blank_template(self, output_path: str) -> str:
        """Boş şablonu indir"""
        if not self.validate_template_exists():
            raise FileNotFoundError(
                f"Şablon eksik: {self.template_path}\n"
                f"Lütfen 'satın alma talep formu.xlsx' dosyasını assets/ klasörüne koyun."
            )
        
        # Template dosyasını kopyala
        import shutil
        shutil.copy2(self.template_path, output_path)
        return output_path
    
    def parse_filled_form(self, file_path: str) -> PurchaseRequest:
        """Dolu formu parse et ve PurchaseRequest oluştur"""
        if not openpyxl:
            raise ImportError("openpyxl kütüphanesi yüklü değil. pip install openpyxl")
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Dosya bulunamadı: {file_path}")
        
        # Excel dosyasını yükle
        workbook = load_workbook(file_path)
        
        # SATINALMA sheet'ini al
        if FORM_MAP["sheet_name"] not in workbook.sheetnames:
            raise ValueError(f"'{FORM_MAP['sheet_name']}' sheet'i bulunamadı")
        
        worksheet = workbook[FORM_MAP["sheet_name"]]
        
        # Meta bilgileri oku
        meta_info = self._parse_meta_info(worksheet)
        
        # Kalemleri oku
        items = self._parse_items(worksheet)
        
        # PurchaseRequest oluştur
        purchase_request = PurchaseRequest(
            talep_no=meta_info.get("talep_no", ""),
            stf_no=meta_info.get("stf_no", ""),
            stf_tarihi=meta_info.get("stf_tarihi", ""),
            santiye=meta_info.get("santiye", ""),
            alim_yeri=meta_info.get("alim_yeri", ""),
            requester=meta_info.get("requester", "Bilinmeyen"),
            items=items,
            notes=meta_info.get("notes", ""),
            status="draft"
        )
        
        workbook.close()
        return purchase_request
    
    def _parse_meta_info(self, worksheet) -> Dict[str, str]:
        """Meta bilgileri parse et"""
        meta_info = {}
        
        # Her meta satırı için değer oku
        for field, row_num in FORM_MAP["meta_rows"].items():
            # İlgili kolondan başlayarak sağa doğru ilk boş olmayan değeri bul
            value = self._find_meta_value(worksheet, row_num, FORM_MAP["meta_value_search_right_from_col"])
            meta_info[field] = value or ""
        
        # Talep eden bilgisini oku (row 13)
        requester_value = self._find_meta_value(worksheet, 13, FORM_MAP["meta_value_search_right_from_col"])
        meta_info["requester"] = requester_value or "Bilinmeyen"
        
        # Notlar (row 16)
        notes_value = self._find_meta_value(worksheet, 16, FORM_MAP["meta_value_search_right_from_col"])
        meta_info["notes"] = notes_value or ""
        
        return meta_info
    
    def _find_meta_value(self, worksheet, row_num: int, start_col: int) -> Optional[str]:
        """Meta satırında sağa doğru ilk boş olmayan değeri bul"""
        for col in range(start_col, 14):  # I'dan M'ye kadar (9-13)
            cell_value = worksheet.cell(row=row_num, column=col).value
            if cell_value and str(cell_value).strip():
                return str(cell_value).strip()
        return None
    
    def _parse_items(self, worksheet) -> List[PurchaseItem]:
        """Kalemleri parse et"""
        items = []
        header_row = FORM_MAP["header_row"]
        data_start_row = FORM_MAP["data_start_row"]
        
        # Header kontrolü
        self._validate_headers(worksheet, header_row)
        
        # Veri satırlarını oku
        row = data_start_row
        while True:
            # Sıra No kontrolü - boşsa dur
            sira_no_cell = worksheet[f"{FORM_MAP['columns']['sira_no']}{row}"]
            if not sira_no_cell.value or str(sira_no_cell.value).strip() == "":
                # Tamamen boş satır kontrolü
                if self._is_empty_row(worksheet, row):
                    break
                else:
                    row += 1
                    continue
            
            # Kalem bilgilerini oku
            item = self._parse_single_item(worksheet, row)
            if item:
                items.append(item)
            
            row += 1
        
        return items
    
    def _validate_headers(self, worksheet, header_row: int):
        """Header satırını validate et"""
        expected_headers = {
            "A": "Sıra No.",
            "B": "Malzeme Kodu",
            "C": "Malzeme Tanımı",
            "D": "MARKA",
            "E": "Birimi",
            "F": "Miktarı",
            "G": "Istenilen Teslim Tarihi",
            "H": "Ambardaki Miktar",
            "I": "Sipariş Miktarı",
            "J": "Hedef Fiyat",
            "L": "Genel Toplam TL"
        }
        
        for col, expected_header in expected_headers.items():
            cell_value = worksheet[f"{col}{header_row}"].value
            if not cell_value or str(cell_value).strip() != expected_header:
                raise ValueError(f"Header uyumsuzluğu: {col}{header_row} beklenen '{expected_header}', bulunan '{cell_value}'")
    
    def _parse_single_item(self, worksheet, row: int) -> Optional[PurchaseItem]:
        """Tek bir kalem parse et"""
        try:
            # Sıra No
            sira_no = self._get_cell_value(worksheet, row, "sira_no")
            if not sira_no:
                return None
            
            sira_no = int(float(sira_no)) if sira_no else 0
            
            # Diğer alanlar
            malzeme_kodu = self._get_cell_value(worksheet, row, "malzeme_kodu") or ""
            malzeme_tanimi = self._get_cell_value(worksheet, row, "malzeme_tanimi") or ""
            marka = self._get_cell_value(worksheet, row, "marka") or ""
            birim = self._get_cell_value(worksheet, row, "birim") or ""
            
            # Numeric alanlar
            miktar = self._parse_numeric_value(worksheet, row, "miktar", 0.0)
            ambardaki_miktar = self._parse_numeric_value(worksheet, row, "ambardaki_miktar")
            siparis_miktari = self._parse_numeric_value(worksheet, row, "siparis_miktari")
            hedef_fiyat = self._parse_numeric_value(worksheet, row, "hedef_fiyat")
            genel_toplam_tl = self._parse_numeric_value(worksheet, row, "genel_toplam_tl")
            
            # Tarih alanları
            istenilen_teslim_tarihi = self._parse_date_value(worksheet, row, "istenilen_teslim_tarihi")
            
            return PurchaseItem(
                sira_no=sira_no,
                malzeme_kodu=malzeme_kodu,
                malzeme_tanimi=malzeme_tanimi,
                marka=marka,
                birim=birim,
                miktar=miktar,
                istenilen_teslim_tarihi=istenilen_teslim_tarihi,
                ambardaki_miktar=ambardaki_miktar,
                siparis_miktari=siparis_miktari,
                hedef_fiyat=hedef_fiyat,
                genel_toplam_tl=genel_toplam_tl
            )
        
        except Exception as e:
            print(f"Satır {row} parse edilemedi: {str(e)}")
            return None
    
    def _get_cell_value(self, worksheet, row: int, column_key: str) -> Optional[str]:
        """Hücre değerini string olarak al"""
        col = FORM_MAP["columns"][column_key]
        cell_value = worksheet[f"{col}{row}"].value
        if cell_value is None:
            return None
        return str(cell_value).strip()
    
    def _parse_numeric_value(self, worksheet, row: int, column_key: str, default: Optional[float] = None) -> Optional[float]:
        """Numeric değer parse et"""
        col = FORM_MAP["columns"][column_key]
        cell_value = worksheet[f"{col}{row}"].value
        
        if cell_value is None:
            return default
        
        try:
            # Excel serial number kontrolü
            if isinstance(cell_value, (int, float)):
                return float(cell_value)
            
            # String değeri numeric'e çevir
            numeric_value = pd.to_numeric(str(cell_value).strip(), errors='coerce')
            return float(numeric_value) if not pd.isna(numeric_value) else default
        
        except (ValueError, TypeError):
            return default
    
    def _parse_date_value(self, worksheet, row: int, column_key: str) -> Optional[str]:
        """Tarih değerini parse et"""
        col = FORM_MAP["columns"][column_key]
        cell_value = worksheet[f"{col}{row}"].value
        
        if cell_value is None:
            return None
        
        try:
            # Excel date serial number
            if isinstance(cell_value, (int, float)):
                from openpyxl.utils.datetime import from_excel
                date_obj = from_excel(cell_value)
                return date_obj.strftime("%Y-%m-%d")
            
            # String date
            if isinstance(cell_value, str):
                # Yaygın tarih formatlarını dene
                date_formats = ["%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]
                for fmt in date_formats:
                    try:
                        date_obj = datetime.strptime(cell_value.strip(), fmt)
                        return date_obj.strftime("%Y-%m-%d")
                    except ValueError:
                        continue
            
            # Datetime object
            if isinstance(cell_value, datetime):
                return cell_value.strftime("%Y-%m-%d")
            
            return str(cell_value).strip()
        
        except Exception:
            return None
    
    def _is_empty_row(self, worksheet, row: int) -> bool:
        """Satırın tamamen boş olup olmadığını kontrol et"""
        for col in range(1, 14):  # A'dan M'ye kadar
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value and str(cell_value).strip():
                return False
        return True
    
    def get_template_info(self) -> Dict[str, Any]:
        """Template dosyası hakkında bilgi getir"""
        return {
            "template_path": self.template_path,
            "template_exists": self.validate_template_exists(),
            "form_map": FORM_MAP,
            "supported_formats": ["xlsx"],
            "required_sheet": FORM_MAP["sheet_name"]
        }
