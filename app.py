# app.py — Teklifbul: Excel içe/aktarım + stil-korumalı mukayese & talep formu üretimi
# Yeni istekler:
# - Talep Kodu -> K3
# - Şantiye    -> K1
# - STF No     -> K2
# - Ürün başlıkları 5. satırda, 1. ürün 6. satırdan başlıyor
# - Şablonda 4 satır olsa bile ürün sayısına göre otomatik satır uzatma (stil/kenarlık/merge korunur)

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from copy import copy
import io, re
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

app = FastAPI(title="Teklifbul - Excel Otomasyon")

# ===================== SABİTLER =====================
# (app.py ile AYNI klasörde olmalılar)
PURCHASE_XLSX = "satın alma talep formu.xlsx"  # talep formu şablonu (pano)
COMPARE_XLSX  = "teklif mukayese formu.xlsx"   # mukayese şablonu
BID_XLSX      = "teklif formu.xlsx"            # (opsiyonel) örnek teklif şablonu

# Üst pano alanları (talep formu)
PURCHASE_FORM_MAP = {
    "company_name": "A1",
    "talep_konusu": "B5",
    "stf_no":       "K2",   # ✅ STF No
    "talep_kodu":   "K3",   # ✅ Talep Kodu
    "santiye":      "K1",   # ✅ Şantiye
    "usd_try":      "T4",   # değişmediyse kalsın; yoksa sil
    "termin_tarihi":"B23",
    "alim_yeri":    "B24",
    "aciklama":     "F25",
}

# Ürün satır başlangıcı (hem okuma hem yazma)
START_ROW_ITEMS = 6   # ✅ 1. ürün 6. satırda
EASY_START_ROW  = 6

# Ürün başlıkları 5. satırda; sütun eşleşmeleri
# A5=Sıra No, B5=Malzeme Kodu, C5=Malzeme Tanımı, D5=Marka,
# E5=Miktar, F5=Birim, G5=İstenen Teslim Tarihi, H5=Ambardaki Miktar,
# I5=Sipariş Miktarı, J5=Hedef Fiyat, L5=Genel Toplam
EASY_COLS_DEFAULT: Dict[str, str] = {
    "no":          "A",
    "sku":         "B",
    "name":        "C",
    "brand":       "D",
    "qty":         "E",
    "unit":        "F",
    "req_date":    "G",
    "stock_qty":   "H",  # opsiyonel
    "order_qty":   "I",  # opsiyonel
    "target_price":"J",  # opsiyonel
    "grand_total": "L",  # opsiyonel
}

# Mukayese şablonu üst bilgiler (değişmediyse böyle kalsın)
COMPARE_TOP_MAP = {
    "talep_kodu": "T1",
    "stf_no": "T2",
    "santiye": "T3",
    "usd_try": "T4",
    "talep_konusu": "B5",
}

# Mukayese tablosu satır/sütun ayarları
TOTAL_ROW, PAY_ROW, TERM_ROW, DELIV_ROW, NOTE_ROW = 21, 22, 23, 24, 25
# Firma blok başlangıç sütunları (F/G, I/J, L/M, O/P, R/S ...)
BASE_COLS = ["F", "I", "L", "O", "R"]

# Talep formu şablonunda hazır ürün satırı sayısı (sen 4 satırlı bırakıyorsun: 6,7,8,9)
PURCHASE_TEMPLATE_ROWS = 4


# ===================== MODELLER =====================
class DemandItem(BaseModel):
    no: int
    name: str
    qty: float
    unit: str
    sku: Optional[str] = None
    brand: Optional[str] = None
    req_date: Optional[str] = None

class BidFirm(BaseModel):
    firma_adi: str
    tel: Optional[str] = None
    odeme: Optional[str] = None
    termin: Optional[str] = None
    fiyatlar: List[Optional[float]]

class Demand(BaseModel):
    stf_no: str
    talep_kodu: str
    santiye: Optional[str] = None
    usd_try: Optional[float] = None
    talep_konusu: Optional[str] = None
    alim_yeri: Optional[str] = None
    aciklama: Optional[str] = None
    items: List[DemandItem]
    firms: List[BidFirm] = Field(default_factory=list)


# ===================== YARDIMCILAR =====================
def _clone_sheet_with_styles(src_wb_path: str):
    """Şablonu (kenarlık/renk/merge/sütun&satır ölçüleri) birebir RAM'e kopyala."""
    src_wb = openpyxl.load_workbook(src_wb_path, data_only=False)
    src_ws = src_wb.active
    dst_wb = openpyxl.Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = src_ws.title

    # Hücre değer + stil kopyası
    for row in src_ws.iter_rows():
        for sc in row:
            dc = dst_ws.cell(row=sc.row, column=sc.column, value=sc.value)
            if sc.has_style:
                dc.font = copy(sc.font)
                dc.border = copy(sc.border)
                dc.fill = copy(sc.fill)
                dc.number_format = copy(sc.number_format)
                dc.protection = copy(sc.protection)
                dc.alignment = copy(sc.alignment)

    # Sütun genişlikleri
    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width

    # Satır yükseklikleri
    for idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[idx].height = dim.height

    # Birleştirmeler
    for r in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(r))

    return dst_wb, dst_ws

def _to_float_safe(val) -> float:
    """Boş/metin gelirse 0; '12,5' → 12.5 çevirir."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "")
    if s == "":
        return 0.0
    if re.fullmatch(r"[0-9\.,]+", s):
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except:
            return 0.0
    return 0.0

def _detect_easy_cols(ws: Worksheet, guess_row: int = 5) -> Dict[str, str]:
    """Başlıklar 5. satırda bekleniyor; bulunamazsa varsayılanları kullan."""
    return EASY_COLS_DEFAULT.copy()

def _read_items_easy(ws: Worksheet) -> List[DemandItem]:
    """Satır 6'dan aşağı ürünleri okur; 'Malzeme Tanımı' boşsa durur."""
    cols = _detect_easy_cols(ws, guess_row=5)
    items: List[DemandItem] = []
    r = EASY_START_ROW
    while r <= ws.max_row:
        name = ws[f"{cols['name']}{r}"].value
        if name in (None, "", " "):  # Malzeme Tanımı boşsa bitti
            break

        no    = ws[f"{cols['no']}{r}"].value
        sku   = ws[f"{cols['sku']}{r}"].value if 'sku' in cols else None
        brand = ws[f"{cols['brand']}{r}"].value if 'brand' in cols else None
        qty   = ws[f"{cols['qty']}{r}"].value
        unit  = ws[f"{cols['unit']}{r}"].value
        reqd  = ws[f"{cols['req_date']}{r}"].value if 'req_date' in cols else None

        try_no  = int(no) if isinstance(no,(int,float)) else (len(items)+1)
        try_qty = _to_float_safe(qty)

        items.append(DemandItem(
            no=try_no,
            name=str(name),
            qty=try_qty,
            unit=str(unit or ""),
            sku=(str(sku) if sku not in (None,"") else None),
            brand=(str(brand) if brand not in (None,"") else None),
            req_date=(str(reqd) if reqd not in (None,"") else None),
        ))
        r += 1
    return items

def _copy_row_style(src_ws, dst_ws, src_row: int, dst_row: int):
    """Bir satırın stilini (kenarlık, font, dolgu, numara formatı, hizalama, yükseklik) kopyalar."""
    if src_row in src_ws.row_dimensions:
        dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height

    max_col = src_ws.max_column
    for c in range(1, max_col + 1):
        sc = src_ws.cell(row=src_row, column=c)
        dc = dst_ws.cell(row=dst_row, column=c)
        if sc.has_style:
            dc.font = copy(sc.font)
            dc.border = copy(sc.border)
            dc.fill = copy(sc.fill)
            dc.number_format = copy(sc.number_format)
            dc.protection = copy(sc.protection)
            dc.alignment = copy(sc.alignment)

def _clone_merges_for_row(ws, template_row: int, new_rows: List[int]):
    """Template satırında yatay merge varsa, yeni satırlara da uygular."""
    to_merge = []
    for m in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = m.bounds
        if min_row == max_row == template_row:
            to_merge.append((min_col, max_col))
    for r in new_rows:
        for min_col, max_col in to_merge:
            ws.merge_cells(start_row=r, start_column=min_col, end_row=r, end_column=max_col)

def _ensure_item_rows_with_style(ws, start_row: int, template_rows: int, need_rows: int):
    """
    Şablonda 'template_rows' kadar satır var (örn. 4). 'need_rows' kadar satıra çıkar.
    Yeni satırlara referans satırın stilini uygula, yatay birleşmeleri kopyala.
    """
    if need_rows <= template_rows:
        return
    extra = need_rows - template_rows
    insert_at = start_row + template_rows          # örn. 6+4=10 -> 10'dan itibaren ekle
    ws.insert_rows(insert_at, amount=extra)

    template_row = start_row + template_rows - 1   # örn. 9. satır referans alınır
    new_rows = list(range(insert_at, insert_at + extra))
    for r in new_rows:
        _copy_row_style(ws, ws, template_row, r)
    _clone_merges_for_row(ws, template_row, new_rows)

def _ensure_col(fi:int) -> str:
    if fi < len(BASE_COLS):
        return BASE_COLS[fi]
    return BASE_COLS[-1]  # şablonda fazlası yoksa son blokta kal


# ===================== ENDPOINTS =====================

@app.post("/import-purchase-form-easy")
async def import_purchase_form_easy(file: UploadFile = File(...)):
    """
    Satın alma formunu (pano) Excel'den OKUR → Demand JSON üretir.
    - Üst bilgiler: PURCHASE_FORM_MAP adreslerinden
    - Ürünler: 6. satırdan başlayarak (A=No, B=Kod, C=Tanım, D=Marka, E=Miktar, F=Birim, G=Termin)
    """
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # Üst alanlar
        top: Dict[str, Any] = {}
        for k, addr in PURCHASE_FORM_MAP.items():
            try:
                top[k] = ws[addr].value
            except Exception:
                top[k] = None

        # Ürünler
        items = _read_items_easy(ws)

        demand = Demand(
            stf_no=str(top.get("stf_no") or ""),
            talep_kodu=str(top.get("talep_kodu") or ""),
            santiye=top.get("santiye"),
            usd_try=_to_float_safe(top.get("usd_try")),
            talep_konusu=top.get("talep_konusu"),
            alim_yeri=top.get("alim_yeri"),
            aciklama=top.get("aciklama"),
            items=items,
            firms=[],
        )
        return JSONResponse(demand.model_dump())
    except Exception as e:
        raise HTTPException(400, f"Kolay mod içe aktarma hatası: {e}")

@app.post("/export-purchase-form")
async def export_purchase_form(demand: Demand):
    """
    Satın alma talep formu Excel'i ÜRETİR (şablon: 'satın alma talep formu.xlsx'):
    - Talep Kodu K3, Şantiye K1, STF No K2 dahil üst bilgiler yazılır
    - 1. ürün 6. satırdan başlar; ürün sayısı kadar satır AŞAĞI DOĞRU UZATILIR (stil/kenarlık korunur)
    - Ürün kolonları: A=No, B=Kod, C=Tanım, D=Marka, E=Miktar, F=Birim, G=İstenen Teslim Tarihi
    """
    try:
        wb, ws = _clone_sheet_with_styles(PURCHASE_XLSX)

        # Üst alanları yaz
        for k, addr in PURCHASE_FORM_MAP.items():
            val = getattr(demand, k, None)
            if val is not None:
                ws[addr].value = val

        # Satır uzatma
        need_rows = max(len(demand.items), PURCHASE_TEMPLATE_ROWS)
        _ensure_item_rows_with_style(
            ws, start_row=START_ROW_ITEMS,
            template_rows=PURCHASE_TEMPLATE_ROWS,
            need_rows=need_rows
        )

        # Ürünleri yaz
        row = START_ROW_ITEMS
        for i, it in enumerate(demand.items, start=0):
            r = row + i
            ws[f"A{r}"].value = it.no if isinstance(it.no, int) else (i + 1)  # Sıra No
            ws[f"B{r}"].value = it.sku or ""                                  # Malzeme Kodu
            ws[f"C{r}"].value = it.name                                       # Malzeme Tanımı
            ws[f"D{r}"].value = it.brand or ""                                # Marka
            ws[f"E{r}"].value = it.qty                                        # Miktar
            ws[f"F{r}"].value = it.unit                                       # Birim
            ws[f"G{r}"].value = it.req_date or ""                             # İstenen Teslim Tarihi
            # H/I/J/L sahalarını istersen burada doldurabilirsin

        out = io.BytesIO()
        name = f"satinalma_talep_{(demand.talep_kodu or demand.stf_no or 'TALEP')}.xlsx"
        wb.save(out); out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={name}"}
        )
    except Exception as e:
        raise HTTPException(400, f"Satın alma formu oluşturulamadı: {e}")

@app.post("/import-bid-form")
async def import_bid_form(file: UploadFile = File(...)):
    """
    (Opsiyonel) Teklif Excel'ini OKUR → {firma_adi, fiyatlar[], items[]} döner.
    Ürünleri talep şablonundaki kolonlardan okur (A..G), 'Birim Fiyat' için F sütunundan sonrası tahmin yapar.
    """
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # Ürünler
        items = _read_items_easy(ws)

        # Birim fiyat başlığı bulmaya çalışma (5. satırda "birim fiyat" arar), bulunamazsa J ya da K tahmini
        price_col = None
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(5, c).value or "").strip().lower()
            if "birim fiyat" in v or "fiyat" in v:
                price_col = c
                break
        if not price_col:
            # J (hedef fiyat) veya K varsayımları
            guess = ws.max_column
            price_col = max(column_index_from_string("J"), min(guess, column_index_from_string("K")))

        # fiyatlar[]
        fiyatlar: List[Optional[float]] = []
        r = EASY_START_ROW
        for _ in items:
            val = ws.cell(r, price_col).value
            fiyatlar.append(_to_float_safe(val) if val not in (None, "") else None)
            r += 1

        out = {
            "firma_adi": (ws["A1"].value or ws["B1"].value or ""),
            "fiyatlar": fiyatlar,
            "items": [i.model_dump() for i in items],
        }
        return JSONResponse(out)
    except Exception as e:
        raise HTTPException(400, f"Teklif formu içe aktarma hatası: {e}")

@app.post("/export-comparison")
async def export_comparison(demand: Demand):
    """
    Mukayese Excel'i ÜRETİR (şablon: 'teklif mukayese formu.xlsx'):
    - Üst alanları doldurur (COMPARE_TOP_MAP)
    - Ürünleri B/C/D/E'ye sırasıyla yazar
    - Firmaları toplam tutara göre sıralar; fiyatları F/I/L/... bloklarına yazar
    - Kenarlık/biçim/merge'ler korunur
    """
    try:
        wb, ws = _clone_sheet_with_styles(COMPARE_XLSX)

        # Üst alanlar
        for k, addr in COMPARE_TOP_MAP.items():
            val = getattr(demand, k, None)
            if val is not None:
                ws[addr].value = val

        # Ürünler (B=No, C=Ürün, D=Miktar, E=Birim — mukayese şablonunun düzenine göre)
        for i, it in enumerate(demand.items):
            r = 8 + i  # mukayese şablonunda ürünler 8. satırdan başlıyorsa
            ws[f"B{r}"].value = it.no
            ws[f"C{r}"].value = it.name
            ws[f"D{r}"].value = it.qty
            ws[f"E{r}"].value = it.unit

        # Firmaları toplam fiyata göre sırala (en ucuz solda)
        def firm_total(f: BidFirm) -> float:
            t = 0.0
            for i2, it2 in enumerate(demand.items):
                p = f.fiyatlar[i2] if i2 < len(f.fiyatlar) else None
                if p is not None:
                    t += float(p) * (it2.qty or 0)
            return t

        firms_sorted = sorted(
            [BidFirm(**(f.model_dump() if isinstance(f, BidFirm) else f)) for f in demand.firms],
            key=firm_total
        )

        for fi, firm in enumerate(firms_sorted):
            if fi >= len(BASE_COLS):
                break  # şablonda kaç blok varsa o kadar
            col = BASE_COLS[fi]
            cidx = column_index_from_string(col)

            ws[f"{col}6"].value = f"{firm.firma_adi} {('('+firm.tel+')') if firm.tel else ''}"

            grand = 0.0
            for i2, it2 in enumerate(demand.items):
                r = 8 + i2
                if i2 < len(firm.fiyatlar) and firm.fiyatlar[i2] is not None:
                    price = float(firm.fiyatlar[i2])
                    ws[f"{col}{r}"].value = price
                    next_col = get_column_letter(cidx + 1)  # yanındaki TOPLAM hücresi
                    line_total = price * (it2.qty or 0)
                    ws[f"{next_col}{r}"].value = line_total
                    grand += line_total

            ws[f"{col}{TOTAL_ROW}"].value = grand
            ws[f"{col}{PAY_ROW}"].value   = firm.odeme or ""
            ws[f"{col}{TERM_ROW}"].value  = firm.termin or ""
            ws[f"{col}{DELIV_ROW}"].value = demand.alim_yeri or ""
            ws[f"{col}{NOTE_ROW}"].value  = demand.aciklama or ""

        out = io.BytesIO()
        name = f"mukayese_tablo_{(demand.stf_no or 'STF')}.xlsx"
        wb.save(out); out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={name}"}
        )
    except Exception as e:
        raise HTTPException(400, f"Mukayese Excel'i oluşturulamadı: {e}")
