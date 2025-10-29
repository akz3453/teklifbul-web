"""
Satın Alma Talep Formu İçe Aktarma (Excel -> JSON)
C:\Users\faruk\OneDrive\Desktop\teklifbul-web\assets\satın alma talep formu konumundaki tablo.

Amaç:
- 'satın alma talep formu.xlsx' şablonunu oku.
- Üst bilgiler ve ürün kalemlerini eşleştir.
- Çıktıyı talep.json olarak kaydet (ve terminale yaz).

Gereksinim:
  pip install openpyxl

Not:
- Şablon hücreleri değişirse sadece PURCHASE_FORM_MAP ve/veya ITEM_HEADER_KEYS'i güncelle.
- Dosyayı proje köküne koy: ./satın alma talep formu.xlsx
"""

import json
import sys
import io
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# --------- ŞABLON HÜCRE EŞLEŞTİRME (ÜST BİLGİ) ---------
PURCHASE_FORM_MAP = {
    "company_name": "A1",    # Talep eden firma ismi (biçim bozulmaz)
    "talep_konusu": "B5",    # Talep Konusu (sistemde yeni alan)
    "stf_no": "S2",          # STF No
    "talep_kodu": "T1",      # Talep Kodu (yanı S1'de yazıyorsa buna göre değiştir)
    "santiye": "S3",         # Şantiye
    "usd_try": "T4",         # TCMB USD/TRY
    "termin_tarihi": "B23",  # Termin Tarihi *
    "alim_yeri": "B24",      # Teslim Şekli / Alım Yeri
    "aciklama": "F25",       # Açıklama
}

# --------- ÜRÜN TABLOSU BAŞLIK ANAHTARLARI ---------
ITEM_HEADER_KEYS = [
    ("no",      ["no", "sıra no", "sıra"]),
    ("sku",     ["malzeme kodu", "stok kodu", "sku", "kod"]),
    ("name",    ["malzeme tanımı", "tanım", "ürün adı", "malzeme"]),
    ("brand",   ["marka/model", "marka", "model"]),
    ("qty",     ["miktar", "adet"]),
    ("unit",    ["birim"]),
    ("req_date",["istenen teslim tarihi", "termin"]),
]

def _find_header_row(ws: Worksheet) -> Dict[str, int]:
    """Başlıkları metinden bul ve kolon indeks haritası döndür."""
    norm = lambda v: str(v).strip().lower() if v is not None else ""
    for r in range(1, ws.max_row + 1):
        row_vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        hits = {}
        for key, aliases in ITEM_HEADER_KEYS:
            for a in aliases:
                if a in row_vals:
                    hits[key] = row_vals.index(a) + 1
                    break
        if len(hits) >= 5:
            return {k: hits.get(k) for k, _ in ITEM_HEADER_KEYS}
    raise RuntimeError("Ürün başlık satırı bulunamadı. Başlık metinlerini kontrol edin.")

def read_top(ws: Worksheet) -> Dict[str, Any]:
    data = {}
    for k, addr in PURCHASE_FORM_MAP.items():
        try:
            data[k] = ws[addr].value
        except Exception:
            data[k] = None
    return data

def read_items(ws: Worksheet, header_map: Dict[str,int]) -> List[Dict[str,Any]]:
    items = []
    # başlık satırından hemen sonraki satırı bul
    start_row = None
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, header_map["no"])
        if str(cell.value).strip().lower() in ["no", "sıra no", "sıra"]:
            start_row = r + 1
            break
    if start_row is None:
        start_row = max(2, 1)

    norm = lambda v: None if v is None else (str(v).strip() if isinstance(v, str) else v)

    r = start_row
    while r <= ws.max_row:
        v_no   = ws.cell(r, header_map["no"]).value if header_map.get("no") else None
        v_name = ws.cell(r, header_map["name"]).value if header_map.get("name") else None
        # Boş satırda dur
        if v_no in (None, "", "-") and (v_name in (None, "")):
            break
        item = {
            "no": int(v_no) if (isinstance(v_no, (int,float)) and v_no == v_no) else (len(items)+1),
            "sku": norm(ws.cell(r, header_map["sku"]).value) if header_map.get("sku") else None,
            "name": norm(v_name) or "",
            "brand": norm(ws.cell(r, header_map["brand"]).value) if header_map.get("brand") else None,
            "qty": float(ws.cell(r, header_map["qty"]).value or 0) if header_map.get("qty") else 0,
            "unit": norm(ws.cell(r, header_map["unit"]).value) if header_map.get("unit") else "",
            "req_date": norm(ws.cell(r, header_map["req_date"]).value) if header_map.get("req_date") else None,
        }
        items.append(item)
        r += 1
    return items

def main(xlsx_path="assets/satın alma talep formu.xlsx", out_json="talep.json"):
    try:
        print(f"📊 Excel dosyası okunuyor: {xlsx_path}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        print("📋 Üst bilgiler okunuyor...")
        top = read_top(ws)
        
        print("🔍 Ürün başlıkları aranıyor...")
        header_map = _find_header_row(ws)
        print(f"✅ Başlık haritası bulundu: {header_map}")
        
        print("📦 Ürün kalemleri okunuyor...")
        items = read_items(ws, header_map)
        print(f"✅ {len(items)} adet ürün kalemi bulundu")

        payload = {
            "company_name": top.get("company_name"),
            "talep_konusu": top.get("talep_konusu"),
            "stf_no": str(top.get("stf_no") or ""),
            "talep_kodu": str(top.get("talep_kodu") or ""),
            "santiye": top.get("santiye"),
            "usd_try": top.get("usd_try"),
            "termin_tarihi": top.get("termin_tarihi"),
            "alim_yeri": top.get("alim_yeri"),
            "aciklama": top.get("aciklama"),
            "items": items,
        }

        print(f"💾 JSON dosyası kaydediliyor: {out_json}")
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        print("\n✅ talep.json oluşturuldu:\n")
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        
    except FileNotFoundError:
        print(f"❌ Excel dosyası bulunamadı: {xlsx_path}")
        print("📁 Dosyayı proje köküne koyun: ./satın alma talep formu.xlsx")
    except Exception as e:
        print(f"❌ Hata oluştu: {e}")
        print("🔧 Şablon hücrelerini kontrol edin veya PURCHASE_FORM_MAP'i güncelleyin")

if __name__ == "__main__":
    # python import_purchase.py  (aynı klasörde Excel varsa)
    # İstersen: python import_purchase.py "C:/path/satın alma talep formu.xlsx"
    if len(sys.argv) > 1:
        main(sys.argv[1])
    else:
        main()
